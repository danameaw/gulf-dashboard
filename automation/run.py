# ── run.py ── Gulf Dashboard Weekly Automation ────────────────────────────
# Schedule: every Wednesday via Windows Task Scheduler
# Usage: python run.py  (or python run.py --week 25 --year 2026)
import sys, os, re, json, glob, shutil, subprocess, argparse, time
from datetime import datetime, timedelta
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(__file__))
from config import (BASE_REPORT_PATH, DASHBOARD_HTML, EXCEL_DIR,
                    GIT_REPO, DASHBOARD_URL,
                    FOLDER_PROJECTS, PROJECT_KEYWORDS, PROJECT_NAMES,
                    MANUAL_OVERRIDES, REPORT_CADENCE, DEFAULT_CADENCE,
                    CADENCE_GRACE_WEEKS)
from extract import extract_from_pdf
from validate import (format_findings, load_history,
                      project_history,
                      record_run, save_history, validate_run, FAIL)


# ── 1. Find latest week folder ─────────────────────────────────────────────
def _parse_folder_date(name):
    """Folder name is 'WEEK_YYMMDD' (e.g. '28_260715') — return the date
    part as an ISO string ('2026-07-15'), or None if it doesn't parse."""
    parts = name.split('_')
    if len(parts) > 1 and len(parts[1]) >= 6:
        try:
            return (f"{2000 + int(parts[1][:2])}-"
                    f"{int(parts[1][2:4]):02d}-{int(parts[1][4:6]):02d}")
        except ValueError:
            pass
    return None


def find_week_folder(year=None):
    pattern = os.path.join(BASE_REPORT_PATH, r"[0-9]*_[0-9]*")
    folders = sorted(glob.glob(pattern))
    if not folders:
        raise FileNotFoundError(f"No week folders found in {BASE_REPORT_PATH}")
    latest = folders[-1]
    name   = os.path.basename(latest)           # e.g. "25_260701"
    parts  = name.split('_')
    week   = int(parts[0])
    yr     = 2000 + int(parts[1][:2]) if len(parts) > 1 else datetime.now().year
    return latest, week, yr, _parse_folder_date(name)


def _norm_key(name):
    """Reduce a folder name to letters+digits for tolerant comparison."""
    return re.sub(r'[^a-z0-9]', '', name.lower())


def resolve_group_folder(week_folder, folder_name):
    """
    Locate a project group's folder inside the week folder.
    These folders are named by hand and the spelling drifts week to week:
    'iWTE' / 'IWTE' / 'Iwte' (harmless, Windows paths are case-insensitive)
    but also 'Wind2027' vs 'Wind 2027' (as in 33_260819), which an exact
    os.path.join misses — reporting FOLDER NOT FOUND for all five Wind
    projects even though the reports were sitting right there. Compare on
    letters and digits only so case, spacing and punctuation don't matter.
    Returns the real path, or None if no folder matches.
    """
    exact = os.path.join(week_folder, folder_name)
    if os.path.isdir(exact):
        return exact
    want = _norm_key(folder_name)
    try:
        for entry in sorted(os.listdir(week_folder)):
            cand = os.path.join(week_folder, entry)
            if os.path.isdir(cand) and _norm_key(entry) == want:
                return cand
    except OSError:
        pass
    return None


# ── 2. Find PDF for a project ──────────────────────────────────────────────
def find_pdf(folder_path, prj_id):
    keywords = PROJECT_KEYWORDS.get(prj_id, [])
    pdfs = glob.glob(os.path.join(folder_path, '*.pdf'))
    pdfs += glob.glob(os.path.join(folder_path, '**', '*.pdf'), recursive=True)
    matches = [pdf for pdf in pdfs
               if any(kw.upper() in os.path.basename(pdf).upper() for kw in keywords)]
    if not matches:
        return None
    # iWTE's regular weekly construction report is always named
    # "Project_<ID>_Construction_weekly_report_...". A one-off report (e.g.
    # a special "Back energize" report covering multiple projects) can
    # legitimately contain a project's keyword too and was winning just by
    # sorting first — prefer the canonically-named file when one exists.
    canonical = [pdf for pdf in matches
                 if os.path.basename(pdf).upper().startswith('PROJECT_')]
    return canonical[0] if canonical else matches[0]


def is_report_due(history, prj_id, week, year):
    """
    Should this project have a report this week?
    Returns (due, note). A weekly project is always due. A monthly or
    biweekly one is due only once its grace period since the last report it
    actually filed has elapsed, so it is not chased every week in between.
    A project with no history at all is treated as due - better to ask than
    to silently stop expecting a report.
    """
    cadence = REPORT_CADENCE.get(prj_id, DEFAULT_CADENCE)
    if cadence == DEFAULT_CADENCE:
        return True, None

    rows = project_history(history, prj_id, before=(year, week))
    if not rows:
        return True, f'{cadence}, no prior report on record'

    last = rows[-1]['week_key']
    m = re.search(r'_W(\d+)_(\d+)$', last)
    if not m:
        return True, cadence
    last_week, last_year = int(m.group(1)), int(m.group(2))
    elapsed = (year - last_year) * 52 + (week - last_week)
    grace = CADENCE_GRACE_WEEKS.get(cadence, 1)
    if elapsed < grace:
        return False, (f'{cadence}, last reported W{last_week}/{last_year} '
                       f'({elapsed} week(s) ago)')
    return True, (f'{cadence}, {elapsed} week(s) since W{last_week}/{last_year}')


# ── 3. Build JS seed snippet ───────────────────────────────────────────────
def js_escape(s):
    return s.replace('\\', '\\\\').replace("'", "\\'").replace('\n', ' ')

def _js_num(v):
    """Format a number (or null) for JS."""
    return 'null' if v is None else str(round(v, 2))

def _build_disciplines_js(discs, indent='    '):
    """Render disciplines dict as JS object literal."""
    if not discs:
        return '{}'
    parts = []
    for disc, vals in discs.items():
        p = _js_num(vals.get('plan'))
        a = _js_num(vals.get('actual'))
        # Weight factor, where the report states one (GMTP's area tables do).
        wf = vals.get('wf')
        wf_js = f", wf: {_js_num(wf)}" if wf is not None else ''
        parts.append(
            f"{indent}  '{disc}': {{ plan: {p}, actual: {a}{wf_js} }}")
    return '{\n' + ',\n'.join(parts) + f'\n{indent}}}'

def _build_scopes_js(scopes):
    """Render multi-scope dict as JS object literal."""
    if not scopes:
        return '{}'
    parts = []
    for scope_name, sc in scopes.items():
        p = _js_num(sc.get('plan'))
        a = _js_num(sc.get('actual'))
        discs = sc.get('disciplines', {})
        if discs:
            disc_js = _build_disciplines_js(discs, indent='      ')
            parts.append(
                f"    '{js_escape(scope_name)}': {{\n"
                f"      plan: {p}, actual: {a},\n"
                f"      disciplines: {disc_js}\n"
                f"    }}"
            )
        else:
            parts.append(
                f"    '{js_escape(scope_name)}': {{ plan: {p}, actual: {a} }}"
            )
    return '{\n' + ',\n'.join(parts) + '\n  }'

def build_seed(prj_id, week, year, data):
    concerns    = ',\n    '.join(f"'{js_escape(c)}'" for c in data['concerns'])
    activities  = ',\n    '.join(f"'{js_escape(a)}'" for a in data['activities'])
    plan        = _js_num(data.get('plan'))
    actual      = _js_num(data.get('actual'))

    # Scopes (Solar/Wind) take priority over flat disciplines (iWTE/GMTP)
    scopes = data.get('scopes', {})
    discs  = data.get('disciplines', {})

    # A project can have both: GMTP reports per-area progress (LNG Tank /
    # BOP & Utility / Marine) *and* a project-wide EPCC breakdown. Emitting
    # only `scopes` dropped the disciplines from the dashboard entirely.
    structure_line = ''
    if scopes:
        structure_line += f"  scopes: {_build_scopes_js(scopes)},\n"
    if discs:
        structure_line += f"  disciplines: {_build_disciplines_js(discs)},\n"

    return (
        f"// ── {PROJECT_NAMES.get(prj_id, prj_id)} ({prj_id}) Week {week} — auto-extracted\n"
        f"seedIfEmpty('{prj_id}_W{week}_{year}', {{\n"
        f"  plan: {plan}, actual: {actual},\n"
        f"{structure_line}"
        f"  concerns: [{concerns}],\n"
        f"  activities: [{activities}],\n"
        f"}});\n"
    )


# ── 4. Update index.html ───────────────────────────────────────────────────
SEED_MARKER          = '// ── SEED DATA (pre-loaded from PDF reports) ──'
MISSING_MARKER_START = '// ── AUTO: MISSING REPORTS ──'
MISSING_MARKER_END   = '// ── END MISSING REPORTS ──'
DATE_MARKER_START    = '// ── AUTO: REPORT DATES ──'
DATE_MARKER_END      = '// ── END REPORT DATES ──'
WEEK_MARKER_START    = '// ── AUTO: CURRENT WEEK ──'
WEEK_MARKER_END      = '// ── END CURRENT WEEK ──'

def _find_seed_inject_point(html):
    """
    Offset just past seedIfEmpty()'s closing brace, where generated seed
    blocks are injected. Anchored on the closing brace rather than the
    function body so edits inside seedIfEmpty can't silently break the match.
    Returns -1 when the marker or the function can't be found.
    """
    marker_pos = html.find(SEED_MARKER)
    if marker_pos == -1:
        return -1
    fn_pos = html.find('function seedIfEmpty(key, data) {', marker_pos)
    if fn_pos == -1:
        return -1
    close_pos = html.find('\n}\n', fn_pos)
    if close_pos == -1:
        return -1
    return close_pos + len('\n}\n')


def _week_seed_header(week, year):
    return f'// ── Week {week}/{year} — AUTO-GENERATED ──'


def _drop_week_seed_block(html, week_header):
    """
    Remove any existing seed block for one week: from its header up to the
    next week header (or the end of the seed region). Re-running a week used
    to insert a SECOND block for the same keys instead of replacing the
    first; the newest block does win (it is injected at the top, and
    seedIfEmpty skips keys already set), but every stale copy stayed in the
    file forever - index.html is already 1.5MB - and made "what is actually
    seeded for W32" impossible to answer by reading the file. Dropping the
    old block first makes a corrective re-run idempotent.
    """
    while True:
        start = html.find(week_header)
        if start == -1:
            return html
        end = html.find('// ── Week ', start + len(week_header))
        if end == -1:
            # Blocks are injected newest-first, so the week being re-run
            # normally has an older week's header after it. Without one
            # this is the oldest block in the file and the only stop point
            # left is </script> - everything between would go with it.
            # Leave it alone and fall back to the old append behaviour.
            print(f'  [warn] no block after {week_header.strip()} - '
                  'leaving the existing one in place')
            return html
        html = html[:start] + html[end:]


def bump_seed_version(html):
    """
    seedIfEmpty() refuses to overwrite a key that already holds real
    plan/actual numbers in a visitor's browser, so re-extracting a week only
    reaches people who have never opened the dashboard - everyone else keeps
    seeing the old, wrong numbers forever. index.html carries a SEED_VERSION
    force-refresh guard for exactly this case, but nothing ever bumped it: it
    still read '2026-07-22.2' while five weekly runs shipped on top of it.
    Stamp it on every run so a correction always reaches every visitor.

    This drops each visitor's cached progressData wholesale, manual dashboard
    edits included, for any week whose seed block does not carry them.
    """
    m = re.search(r"const SEED_VERSION = '([^']*)';", html)
    if not m:
        print("  [warn] SEED_VERSION not found - visitors may keep stale data")
        return html
    stamp = datetime.now().strftime('%Y-%m-%d.%H%M')
    print(f"  ✓ SEED_VERSION {m.group(1)} → {stamp} (forces client refresh)")
    return html[:m.start()] + f"const SEED_VERSION = '{stamp}';" + html[m.end():]


def update_html(week, year, seeds_js, missing_ids, report_date=None):
    with open(DASHBOARD_HTML, encoding='utf-8') as f:
        html = f.read()

    # Inject seed data right after the seedIfEmpty() function that follows
    # SEED_MARKER — anchored on the function's closing brace, not its body.
    # (A previous version matched a hardcoded copy of the function body; once
    # seedIfEmpty's internals were edited in index.html the match silently
    # stopped working and every run fell back to appending a duplicate block
    # at the end of the file. Because seedIfEmpty only overwrites a key that
    # is still null, those appended duplicates could permanently shadow any
    # later correction for a key that already had a non-null value.)
    week_header = _week_seed_header(week, year)
    html = _drop_week_seed_block(html, week_header)
    inject_point = _find_seed_inject_point(html)
    if inject_point != -1:
        new_seeds = '\n\n' + week_header + '\n' + '\n'.join(seeds_js)
        html = html[:inject_point] + new_seeds + html[inject_point:]
    else:
        print("  [warn] SEED_MARKER not found, appending seeds before </script>")
        html = html.replace('</script>', '\n'.join(seeds_js) + '\n</script>', 1)

    # Update missing reports — merge into MISSING_REPORTS_BY_WEEK dict
    week_key = f"W{week}_{year}"
    by_week_re = re.compile(
        r'(// ── AUTO: MISSING REPORTS ──\n)'
        r'const MISSING_REPORTS_BY_WEEK = (\{.*?\});'
        r'(\n// ── END MISSING REPORTS ──)',
        re.DOTALL)
    m = by_week_re.search(html)
    if m:
        existing = json.loads(m.group(2))
        existing[week_key] = missing_ids
        new_block = (
            f"{MISSING_MARKER_START}\n"
            f"const MISSING_REPORTS_BY_WEEK = {json.dumps(existing, separators=(',', ':'))};\n"
            f"{MISSING_MARKER_END}"
        )
        html = html[:m.start()] + new_block + html[m.end():]
    else:
        new_block = (
            f"{MISSING_MARKER_START}\n"
            f"const MISSING_REPORTS_BY_WEEK = {json.dumps({week_key: missing_ids}, separators=(',', ':'))};\n"
            f"{MISSING_MARKER_END}"
        )
        if MISSING_MARKER_START in html:
            start = html.index(MISSING_MARKER_START)
            end   = html.index(MISSING_MARKER_END) + len(MISSING_MARKER_END)
            html  = html[:start] + new_block + html[end:]
        else:
            html = html.replace('<script>', new_block + '\n<script>', 1)

    # Update real report date — merge into REPORT_DATES_BY_WEEK dict
    if report_date:
        date_re = re.compile(
            r'(// ── AUTO: REPORT DATES ──.*?)'
            r'const REPORT_DATES_BY_WEEK = (\{.*?\});'
            r'(.*?// ── END REPORT DATES ──)',
            re.DOTALL)
        m = date_re.search(html)
        if m:
            existing = json.loads(m.group(2))
            existing[week_key] = report_date
            new_block = (
                m.group(1) +
                f"const REPORT_DATES_BY_WEEK = {json.dumps(existing, separators=(',', ':'))};" +
                m.group(3)
            )
            html = html[:m.start()] + new_block + html[m.end():]
        else:
            print("  [warn] REPORT_DATES_BY_WEEK marker not found, skipping report date")

    # Update current week displayed in dashboard
    week_js = (
        f"{WEEK_MARKER_START}\n"
        f"currentWeek = {week}; currentYear = {year}; saveData();\n"
        f"{WEEK_MARKER_END}"
    )
    if WEEK_MARKER_START in html:
        start = html.index(WEEK_MARKER_START)
        end   = html.index(WEEK_MARKER_END) + len(WEEK_MARKER_END)
        html  = html[:start] + week_js + html[end:]
    else:
        # Replace any existing hardcoded currentWeek line
        import re as _re
        html = _re.sub(
            r'currentWeek\s*=\s*\d+;\s*currentYear\s*=\s*\d+;\s*saveData\(\);',
            week_js, html)

    html = bump_seed_version(html)

    with open(DASHBOARD_HTML, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  ✓ index.html updated (Week {week}/{year})")


# ── 5. Update / create Excel ───────────────────────────────────────────────
def update_excel(week, year, results):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("  [skip] openpyxl not installed")
        return

    xl_path = os.path.join(EXCEL_DIR, f'Gulf_Dashboard_W{week}_{year}.xlsx')

    # Load existing or create new
    if os.path.exists(xl_path):
        wb = openpyxl.load_workbook(xl_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    sheet_name = f'W{week}'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Header
    headers = ['PRJ ID', 'Project Name', 'PDF Found', 'Concerns', 'Activities']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F3864')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ri, (prj_id, info) in enumerate(sorted(results.items()), 2):
        ws.cell(ri, 1, prj_id)
        ws.cell(ri, 2, PROJECT_NAMES.get(prj_id, prj_id))
        ws.cell(ri, 3, 'YES' if info['found'] else 'NO')
        ws.cell(ri, 4, '\n'.join(info['data']['concerns']))
        ws.cell(ri, 5, '\n'.join(info['data']['activities']))
        for ci in range(1, 6):
            ws.cell(ri, ci).alignment = Alignment(wrap_text=True, vertical='top')
        if not info['found']:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = PatternFill('solid', fgColor='FFE0E0')

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60

    wb.save(xl_path)
    print(f"  ✓ Excel saved: {xl_path}")


# ── 6. Git commit & push ───────────────────────────────────────────────────
def git_push(week, year):
    try:
        subprocess.run(['git', 'add', '-A'],
                       cwd=GIT_REPO, check=True, capture_output=True)
        subprocess.run(['git', 'commit', '-m',
                        f'Auto: Week {week}/{year} data extracted from PDFs'],
                       cwd=GIT_REPO, check=True, capture_output=True)
        subprocess.run(['git', 'push'],
                       cwd=GIT_REPO, check=True, capture_output=True)
        print("  ✓ Pushed to GitHub")
    except subprocess.CalledProcessError as e:
        print(f"  [git error] {e.stderr.decode(errors='replace') if e.stderr else e}")


# ── 7. Send Email via Outlook (local desktop app) ─────────────────────────
EMAIL_FROM   = "danaya.th@gulf.co.th"
EMAIL_TO     = ["purachet.am@gulf.co.th", "chalong@gulf.co.th"]

# ── The dashboard's Summary view, mirrored into the email ─────────────────
#
# The email used to open with nothing but a count of updated projects, so
# reading it told you the run had happened but not how anything was doing.
# Everything below is a faithful capture of the dashboard's own Summary
# instead - the same five stat cards renderSummary() draws, and the same
# Project Progress Summary table renderTable() draws, in the same project
# order, with the same colours, badges and delay tags.
#
# It is a mirror, deliberately: the status rule, the delay-tag threshold and
# the row order are copied from index.html rather than reinvented, so the
# email and the dashboard can never disagree. When index.html changes, the
# matching constant here has to change with it.
#
# Outlook renders a subset of HTML/CSS - no flexbox, no grid, no <style>
# blocks, no CSS variables - so the layout below is tables with literal
# inline styles throughout, and the dashboard's :root colours are repeated
# here as hex.

# index.html :root
_C_PRIMARY   = '#1a3c5e'
_C_BORDER    = '#dde3ec'
_C_MUTED     = '#6b7a99'
_C_GREEN     = '#27ae60'
_C_RED       = '#e74c3c'
_C_AHEAD     = '#1abc9c'

# .type-<Type>
_TYPE_BADGE = {
    'Solar':          ('#fff3cd', '#856404'),
    'Wind':           ('#cce5ff', '#004085'),
    'WTE':            ('#d4edda', '#155724'),
    'Gas':            ('#e2d9f3', '#432874'),
    'Hydro':          ('#ffe5d0', '#7d3c00'),
    'DataCenter':     ('#d1f2eb', '#0e6655'),
    'Infrastructure': ('#f0f0f0', '#555555'),
}

# .status-<Status>, plus renderTable()'s status icon
_STATUS_BADGE = {
    'On Track': ('#d4edda', '#155724', '&#10004;'),
    'Delay':    ('#f8d7da', '#721c24', '&#9888;'),
    'Ahead':    ('#d1f2eb', '#0e6655', '&#9650;'),
    'TBD':      ('#f0f0f0', '#777777', '&ndash;'),
}

_PROJECT_ROW = re.compile(
    r'id:"(PRJ-\d+)"\s*,\s*name:"([^"]*)"\s*,\s*type:"([^"]*)"\s*,'
    r'\s*contract:"([^"]*)"')


def _load_projects():
    """
    The dashboard's PROJECTS array, in file order - that order IS the table's
    row order, so reading it keeps the email's rows in the same sequence as
    the dashboard instead of re-sorting into something the reader would have
    to re-map. index.html stays the single source of truth for Type and
    Contract too. Returns [] on any read/parse failure; the caller then falls
    back to PROJECT_NAMES so a miss degrades the table rather than breaking
    the send.
    """
    try:
        with open(DASHBOARD_HTML, encoding='utf-8') as f:
            html = f.read()
    except Exception as e:
        print(f"  [email] could not read project list: {e}")
        return []
    return [{'id': pid, 'name': name, 'type': ptype, 'contract': contract}
            for pid, name, ptype, contract in _PROJECT_ROW.findall(html)]


# getDiscAbbr() / getScopeAbbr() in index.html
_DISC_ABBR = {
    'Engineering': 'ENG', 'Procurement': 'PRO', 'Construction': 'CON',
    'Commissioning': 'COM', 'Con & Com': 'C&C', 'Civil': 'CIV',
    'Electrical': 'EE', 'T&C': 'T&C',
}
_SCOPE_ABBR = [
    (re.compile(r'C&E|GPD|GUE', re.I), 'C&E'),
    (re.compile(r'Sub|Siemens', re.I), 'Sub'),
    (re.compile(r'T/L|SCT|ENCOM|RSS', re.I), 'T/L'),
    (re.compile(r'BOP|PCZ', re.I), 'BOP'),
    (re.compile(r'TSA', re.I), 'TSA'),
    (re.compile(r'CBOP', re.I), 'CBOP'),
]


def _disc_abbr(name):
    return _DISC_ABBR.get(name, name[:3].upper())


def _scope_abbr(name):
    for pattern, abbr in _SCOPE_ABBR:
        if pattern.search(name):
            return abbr
    return re.sub(r'\(.*\)', '', name).strip()[:6]


def _pairs(entry):
    """(plan, actual) when both are present, else None."""
    if not isinstance(entry, dict):
        return None
    p, a = entry.get('plan'), entry.get('actual')
    return None if p is None or a is None else (p, a)


def _build_delay_tags(data):
    """
    buildDelayTags() from index.html: every scope/discipline running behind
    by more than 1%, worst first. This is the dashboard's Overall Progress
    column, and it is what explains a row whose overall variance is positive
    yet whose status reads Delay - one discipline is dragging.

    Note the threshold here is -1%, independent of the +/-2% that decides
    status; that is how the dashboard behaves.
    """
    THRESHOLD = -1
    tags = []

    def add_from_disc(disc, prefix=''):
        for dname, dd in (disc or {}).items():
            pair = _pairs(dd)
            if pair is None:
                continue
            v = pair[1] - pair[0]
            if v < THRESHOLD:
                label = (prefix + ' ' if prefix else '') + _disc_abbr(dname)
                tags.append((label, v))

    scopes = data.get('scopes') or {}
    discs  = data.get('disciplines') or {}
    if scopes:
        for sname, sc in scopes.items():
            if isinstance(sc, dict) and sc.get('disciplines'):
                add_from_disc(sc['disciplines'], _scope_abbr(sname))
            else:
                pair = _pairs(sc)
                if pair and pair[1] - pair[0] < THRESHOLD:
                    tags.append((_scope_abbr(sname), pair[1] - pair[0]))
    elif discs:
        add_from_disc(discs)
    else:
        pair = _pairs(data)
        if pair and pair[1] - pair[0] < THRESHOLD:
            tags.append(('Overall', pair[1] - pair[0]))

    return sorted(tags, key=lambda t: t[1])


def _has_progress_data(data):
    """hasProgressData() from index.html."""
    if data.get('scopes'):
        return True
    if data.get('disciplines'):
        return True
    return data.get('plan') is not None and data.get('actual') is not None


def _iter_variances(data):
    """Every scope/discipline variance calcStatus() would consider."""
    def from_disc(disc):
        for dd in (disc or {}).values():
            pair = _pairs(dd)
            if pair:
                yield pair[1] - pair[0]

    scopes = data.get('scopes') or {}
    if scopes:
        for sc in scopes.values():
            if isinstance(sc, dict) and sc.get('disciplines'):
                yield from from_disc(sc['disciplines'])
            else:
                pair = _pairs(sc)
                if pair:
                    yield pair[1] - pair[0]
    else:
        yield from from_disc(data.get('disciplines'))


def _status_of(data):
    """
    calcStatus() from index.html: any single scope or discipline 2% or more
    behind makes the whole project 'Delay', 2% or more ahead makes it
    'Ahead', and only when there is no breakdown at all does the overall
    plan/actual decide.
    """
    variances = list(_iter_variances(data))
    if variances:
        if any(v <= -2 for v in variances):
            return 'Delay'
        if any(v >= 2 for v in variances):
            return 'Ahead'
    plan, actual = data.get('plan'), data.get('actual')
    if plan is None or actual is None:
        return 'TBD'
    diff = actual - plan
    if diff <= -2:
        return 'Delay'
    if diff >= 2:
        return 'Ahead'
    return 'On Track'


# renderSummary()'s five cards, in the order the strip draws them
_STAT_CARDS = [
    ('Total Projects', _C_PRIMARY, 'Reported this week'),
    ('On Track',       _C_GREEN,   'Within &plusmn;2%'),
    ('Delay',          _C_RED,     '&lt; -2% variance'),
    ('Ahead',          _C_AHEAD,   '&gt; +2% variance'),
    ('TBD / No Data',  _C_MUTED,   'Pending import'),
]


def _summary_strip_html(statuses):
    """The five stat cards, as one table row so Outlook lays them out."""
    counts = {
        'Total Projects': len(statuses),
        'On Track':       statuses.count('On Track'),
        'Delay':          statuses.count('Delay'),
        'Ahead':          statuses.count('Ahead'),
        'TBD / No Data':  statuses.count('TBD'),
    }
    cells = ''.join(
        f"<td width='20%' valign='top' style='padding:0 6px 0 0'>"
        f"<table cellspacing='0' cellpadding='0' style='width:100%;"
        f"background:#ffffff;border:1px solid {_C_BORDER};"
        f"border-radius:10px'><tr><td style='padding:14px 16px'>"
        f"<div style='font-size:11px;color:{_C_MUTED};"
        f"text-transform:uppercase;letter-spacing:0.5px'>{label}</div>"
        f"<div style='font-size:26px;font-weight:700;color:{colour};"
        f"line-height:1;padding:4px 0'>{counts[label]}</div>"
        f"<div style='font-size:11px;color:{_C_MUTED}'>{caption}</div>"
        f"</td></tr></table></td>"
        for label, colour, caption in _STAT_CARDS
    )
    return (f"<table cellspacing='0' cellpadding='0' style='width:100%;"
            f"border-collapse:separate;margin-bottom:22px'><tr>{cells}</tr>"
            f"</table>")


def _progress_cell_html(data, no_report):
    """The dashboard's Overall Progress column: its delay tags, verbatim."""
    if no_report:
        return f"<span style='color:{_C_MUTED};font-size:10px'>No data</span>"
    tags = _build_delay_tags(data)
    if not tags:
        if not _has_progress_data(data):
            return f"<span style='color:{_C_MUTED};font-size:10px'>No data</span>"
        return (f"<span style='color:{_C_GREEN};font-size:10px;"
                f"font-weight:700'>&#10004; All on track</span>")
    return ''.join(
        f"<span style='font-size:10px;font-weight:700;color:#ffffff;"
        f"background:{_C_RED};border-radius:3px;padding:1px 5px;"
        f"margin:0 3px 3px 0;display:inline-block;white-space:nowrap'>"
        f"{label} {v:.1f}%</span>"
        for label, v in tags
    )


def _delayed_projects(projects, results, missing):
    """
    The rows the email actually lists: projects the dashboard marks Delay,
    i.e. those with at least one scope or discipline behind plan. Anything
    on track, ahead, or still awaiting a report is left out of the table -
    the stat cards above it still count the full portfolio, so nothing is
    hidden, it is only moved out of the reader's way.
    """
    out = []
    for p in projects:
        entry = results.get(p['id']) or {}
        if p['id'] in missing or not entry.get('found'):
            continue
        data = entry.get('data') or {}
        if _status_of(data) == 'Delay':
            out.append((p, data))
    return out


def _progress_table_html(projects, results, missing):
    """
    renderTable(): every project, in the dashboard's own row order. A project
    with nothing behind plan still gets its row and reads the way the
    dashboard reads it - "All on track", or the No Report badge when its
    report has not arrived.
    """
    rows = []
    for idx, p in enumerate(projects, start=1):
        pid       = p['id']
        entry     = results.get(pid) or {}
        data      = entry.get('data') or {}
        no_report = pid in missing or not entry.get('found')
        status    = 'TBD' if no_report else _status_of(data)
        s_bg, s_fg, s_icon = _STATUS_BADGE[status]
        t_bg, t_fg = _TYPE_BADGE.get(p['type'], ('#f0f0f0', '#555555'))

        missing_badge = (
            f"<span style='font-size:10px;font-weight:700;color:#ffffff;"
            f"background:#f59e0b;border-radius:3px;padding:1px 6px;"
            f"margin-left:6px;white-space:nowrap'>&#9888; No Report</span>"
        ) if no_report else ''

        rows.append(
            f"<tr style='border-bottom:1px solid {_C_BORDER}'>"
            f"<td style='padding:7px 10px;color:{_C_MUTED};font-size:11px'>{idx}</td>"
            f"<td style='padding:7px 10px;color:{_C_MUTED};font-size:11px;"
            f"white-space:nowrap'>{pid}</td>"
            f"<td style='padding:7px 10px'>"
            f"<span style='font-weight:700;color:{_C_PRIMARY};font-size:12px'>"
            f"{p['name']}</span>{missing_badge}</td>"
            f"<td style='padding:7px 10px'>"
            f"<span style='background:{t_bg};color:{t_fg};font-size:10px;"
            f"font-weight:700;text-transform:uppercase;letter-spacing:0.4px;"
            f"padding:2px 8px;border-radius:20px;white-space:nowrap'>"
            f"{p['type']}</span></td>"
            f"<td style='padding:7px 10px'>"
            f"<span style='background:#fff3e0;color:#e65100;"
            f"border:1px solid #ffcc80;border-radius:4px;font-size:10px;"
            f"padding:1px 6px;font-weight:700;white-space:nowrap'>"
            f"{p['contract']}</span></td>"
            f"<td style='padding:7px 10px'>{_progress_cell_html(data, no_report)}</td>"
            f"<td style='padding:7px 10px;white-space:nowrap'>"
            f"<span style='background:{s_bg};color:{s_fg};font-size:11px;"
            f"font-weight:700;padding:3px 9px;border-radius:20px'>"
            f"{s_icon} {status}</span></td>"
            f"</tr>"
        )

    head = ''.join(
        f"<th style='padding:9px 10px;text-align:left;color:#ffffff;"
        f"font-weight:500;font-size:11px;white-space:nowrap'>{label}</th>"
        for label in ('#', 'Project ID', 'Project Name', 'Type', 'Contract',
                      'Overall Progress', 'Status')
    )
    return (f"<table cellspacing='0' cellpadding='0' style='width:100%;"
            f"border-collapse:collapse;margin-bottom:24px;"
            f"border:1px solid {_C_BORDER};border-radius:6px;"
            f"overflow:hidden'>"
            f"<tr style='background:{_C_PRIMARY}'>{head}</tr>"
            f"{''.join(rows)}</table>")


# The corporate e-mail footer. The logo is optional: drop a PNG at
# automation/assets/gulf_logo.png and it is attached to the message and
# referenced by Content-ID - Outlook strips data: URIs, so a CID attachment
# is the only reliable way to show an image in a sent mail. With no file
# present the footer renders as text alone rather than a broken-image icon.
EMAIL_LOGO = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          'assets', 'gulf_logo.png')
EMAIL_LOGO_CID = 'gulflogo'

SIGNATURE_NAME    = 'Danaya Thataporn'
SIGNATURE_TITLE   = 'Engineer II - Project Control'
SIGNATURE_COMPANY = 'Gulf Development Public Company Limited'
SIGNATURE_ADDRESS = ('87 Wireless Road, M Thai Tower 11th Floor, '
                     'All Seasons Place, Lumpini, Pathumwan,')
SIGNATURE_CONTACT = ('Bangkok 10330, Thailand &nbsp;|&nbsp; Tel: +662080-4458 '
                     '&nbsp;|&nbsp; Fax: +66 2080-4455 &nbsp;|&nbsp; ')
SIGNATURE_WEB     = 'www.gulf.co.th'


def _signature_html():
    """The footer block: optional logo cell on the left, details on the right."""
    logo_cell = (
        f"<td valign='middle' style='padding:0 18px 0 0'>"
        f"<img src='cid:{EMAIL_LOGO_CID}' alt='GULF' height='44' "
        f"style='display:block;border:0'></td>"
    ) if os.path.exists(EMAIL_LOGO) else ''

    return (
        f"<table cellspacing='0' cellpadding='0' style='margin:0;"
        f"border-collapse:collapse'><tr>{logo_cell}"
        f"<td valign='middle' style='font-family:Segoe UI,Arial,sans-serif'>"
        f"<div style='font-size:13px;color:#8c8c8c;padding-bottom:1px'>"
        f"{SIGNATURE_NAME} &nbsp;|&nbsp; {SIGNATURE_TITLE}</div>"
        f"<div style='font-size:13px;font-weight:700;color:#1e2535;"
        f"padding-bottom:1px'>{SIGNATURE_COMPANY}</div>"
        f"<div style='font-size:12px;color:#3f3f3f;line-height:1.5'>"
        f"{SIGNATURE_ADDRESS}<br>{SIGNATURE_CONTACT}"
        f"<a href='https://{SIGNATURE_WEB}' style='color:#3f3f3f'>"
        f"{SIGNATURE_WEB}</a></div>"
        f"</td></tr></table>"
    )


def _fmt_cutoff(report_date):
    """
    '2026-08-26' -> '29 Aug 2026': the data cut-off is the Saturday that
    closes the reporting week, not the date on the week folder.

    The folder is named for the day it was created - a Wednesday in practice -
    while the reports inside it cover Saturday through Friday. So the cut-off
    is the first Saturday on or after the folder date; a folder already dated
    a Saturday keeps that same day rather than jumping a week ahead.

    Falls back to the raw string, then to nothing, rather than guessing.
    """
    if not report_date:
        return ''
    try:
        d = datetime.strptime(report_date, '%Y-%m-%d')
    except (ValueError, TypeError):
        return str(report_date)
    return (d + timedelta(days=(5 - d.weekday()) % 7)).strftime('%d %b %Y')


def _build_html_body(week, year, found, missing, report_date=None):
    projects = _load_projects()
    if not projects:
        projects = [{'id': pid, 'name': PROJECT_NAMES.get(pid, pid),
                     'type': 'Infrastructure', 'contract': '&mdash;'}
                    for pid in sorted(set(found) | set(missing))]

    results  = dict(found)
    for pid in missing:
        results.setdefault(pid, {'found': False, 'data': {}})

    statuses = ['TBD' if (p['id'] in missing
                          or not (results.get(p['id']) or {}).get('found'))
                else _status_of(results[p['id']]['data'])
                for p in projects if p['id'] in results]

    delayed  = _delayed_projects(projects, results, missing)

    missing_names = ', '.join(
        (next((p['name'] for p in projects if p['id'] == pid), pid))
        for pid in missing)
    cutoff = _fmt_cutoff(report_date)
    cutoff_line = (
        f"<p style='margin:0 0 22px;font-size:12px;color:{_C_MUTED}'>"
        f"Data cut-off: <b style='color:#1e2535'>{cutoff}</b> &nbsp;|&nbsp; "
        f"Reporting week {week}/{year}</p>"
    ) if cutoff else (
        f"<p style='margin:0 0 22px;font-size:12px;color:{_C_MUTED}'>"
        f"Reporting week {week}/{year}</p>"
    )

    missing_note = (
        f"<p style='margin:0 0 22px'>The <b>{len(missing)} project(s)</b> marked "
        f"<span style='font-size:10px;font-weight:700;color:#ffffff;"
        f"background:#f59e0b;border-radius:3px;padding:1px 6px'>"
        f"&#9888; No Report</span> above &mdash; <b>{missing_names}</b> &mdash; "
        f"have not yet submitted a weekly progress report. Kindly ensure the "
        f"relevant PDF files are uploaded to the ShareDrive at the earliest "
        f"convenience.</p>"
    ) if missing else ""

    return f"""
<html><body style="font-family:Segoe UI,Arial,sans-serif;font-size:13px;color:#1e2535;line-height:1.7;max-width:900px;margin:0 auto;padding:24px;background:#ffffff">
<p style="margin:0 0 4px">Dear P'Tee and P'Hall,</p>
<p style="margin:0 0 6px">
  Please be informed that the <b>Gulf Engineering Dashboard &mdash; Week {week}/{year}</b>
  has been updated. Below is the dashboard's Summary view as it currently stands.
</p>
{cutoff_line}

{_summary_strip_html(statuses)}

<p style="font-size:13px;font-weight:700;color:{_C_PRIMARY};margin:0 0 10px">Project Progress Summary</p>
{_progress_table_html(projects, results, missing)}
<p style="margin:0 0 22px;font-size:11px;color:{_C_MUTED}">
  Overall Progress lists every scope or discipline running more than 1% behind
  plan, worst first &mdash; so a project can read Delay while its overall figure
  is on or ahead of plan. {len(delayed)} of {len(statuses)} project(s) have at
  least one discipline behind; the rest show as on track or ahead.
</p>

{missing_note}

<table cellspacing="0" cellpadding="0" style="margin:0 0 20px">
  <tr><td style="background:{_C_PRIMARY};border-radius:6px;padding:11px 22px">
    <a href="{DASHBOARD_URL}" style="color:#ffffff;font-weight:600;text-decoration:none;font-size:13px">Open the full dashboard &rarr;</a>
  </td></tr>
</table>
<p style="margin:0 0 20px;font-size:12px;color:{_C_MUTED}">
  Direct link: <a href="{DASHBOARD_URL}" style="color:#2d6a9f">{DASHBOARD_URL}</a>
</p>

<p style="margin:0 0 4px">Best Regards,</p>
<p style="margin:0 0 18px;font-weight:500">{SIGNATURE_NAME}</p>
<div style="border-top:1px solid #dde3ec;padding-top:14px">{_signature_html()}</div>
</body></html>"""

def _ensure_outlook_running(wait_seconds=20):
    """Launch Outlook if it's not already running, and give it time to finish
    starting up. A cold Dispatch() launch from inside the COM call is what
    causes the automation to hang, so we start the process separately first."""
    try:
        running = subprocess.run(
            ['tasklist', '/FI', 'IMAGENAME eq OUTLOOK.EXE'],
            capture_output=True, text=True
        )
        if 'OUTLOOK.EXE' not in running.stdout.upper():
            subprocess.Popen(['outlook.exe'], shell=True)
            time.sleep(wait_seconds)
    except Exception as e:
        print(f"  [email] Could not check/launch Outlook: {e}")


def send_email(week, year, results, missing_ids, timeout=90,
               report_date=None):
    found   = {k: v for k, v in results.items() if v['found']}
    missing = missing_ids

    subject   = (f"[Gulf Dashboard] W{week}/{year} Update — "
                 f"{len(found)} Projects Updated, {len(missing)} Reports Pending")
    html_body = _build_html_body(week, year, found, missing,
                                report_date=report_date)

    _ensure_outlook_running()

    payload = {
        'to': EMAIL_TO,
        'subject': subject,
        'html_body': html_body,
        # Referenced by the footer as cid:<cid>. Omitted when the file is
        # absent, in which case the footer renders as text only.
        'inline_image': ({'path': os.path.abspath(EMAIL_LOGO),
                          'cid': EMAIL_LOGO_CID}
                         if os.path.exists(EMAIL_LOGO) else None),
    }
    payload_path = os.path.join(os.path.dirname(__file__), '_email_payload.json')
    sender_script = os.path.join(os.path.dirname(__file__), '_send_outlook_mail.py')
    with open(payload_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f)

    try:
        # Runs in its own process so a stuck Outlook COM call (e.g. waiting on
        # a dialog) can be killed by the timeout instead of hanging run.py.
        subprocess.run(
            [sys.executable, sender_script, payload_path],
            timeout=timeout, check=True, capture_output=True, text=True,
        )
        print(f"  [OK] Email sent via Outlook to: {', '.join(EMAIL_TO)}")
    except subprocess.TimeoutExpired:
        print(f"  [email error] Outlook did not respond within {timeout}s "
              f"(may be showing a dialog or still starting up) — email not sent")
    except subprocess.CalledProcessError as e:
        print(f"  [email error] {e.stderr.strip() if e.stderr else e}")
    except Exception as e:
        print(f"  [email error] {e}")
    finally:
        if os.path.exists(payload_path):
            os.remove(payload_path)


# ── 8. Windows Notification ────────────────────────────────────────────────
def notify(week, year, found_count, missing_count):
    msg = (f"Week {week}/{year} — {found_count} projects updated"
           + (f", {missing_count} missing reports" if missing_count else ""))
    url = DASHBOARD_URL
    try:
        # Try win10toast first
        from win10toast import ToastNotifier
        ToastNotifier().show_toast(
            "Gulf Dashboard Updated", msg,
            duration=10, threaded=True)
    except Exception:
        pass
    # Also open browser to dashboard
    try:
        import webbrowser
        webbrowser.open(url)
    except Exception:
        pass
    # Fallback: PowerShell balloon
    try:
        ps_cmd = (
            f'Add-Type -AssemblyName System.Windows.Forms;'
            f'$n = New-Object System.Windows.Forms.NotifyIcon;'
            f'$n.Icon = [System.Drawing.SystemIcons]::Information;'
            f'$n.Visible = $true;'
            f'$n.ShowBalloonTip(8000,"Gulf Dashboard Updated","{msg}",'
            f'[System.Windows.Forms.ToolTipIcon]::Info);'
            f'Start-Sleep 9; $n.Dispose()'
        )
        subprocess.Popen(['powershell', '-Command', ps_cmd])
    except Exception:
        pass


# ── MAIN ───────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--week', type=int, default=None)
    parser.add_argument('--year', type=int, default=None)
    parser.add_argument('--dry-run', action='store_true',
                        help='Extract only, do not update files')
    parser.add_argument('--no-email', action='store_true',
                        help='Skip sending email notification')
    parser.add_argument('--dump', metavar='PATH', default=None,
                        help='Write the raw extraction results to a JSON file '
                             '(a full run takes minutes, so this lets the '
                             'validation rules be re-checked without one)')
    parser.add_argument('--strict', action='store_true',
                        help='Stop before updating any file if the data '
                             'checks report a FAIL')
    parser.add_argument('--no-push', action='store_true',
                        help='Update index.html/Excel locally but do not '
                             'commit or push (for reviewing a correction '
                             'before it goes live)')
    args = parser.parse_args()

    print("=" * 60)
    print("Gulf Dashboard Weekly Automation")
    print("=" * 60)

    # Find week folder
    if args.week and args.year:
        # Manual override: find folder matching week number
        pattern = os.path.join(BASE_REPORT_PATH, f"{args.week:02d}_*")
        matches = glob.glob(pattern)
        if not matches:
            print(f"  [error] No folder found for week {args.week}")
            sys.exit(1)
        week_folder = matches[0]
        week, year = args.week, args.year
        report_date = _parse_folder_date(os.path.basename(week_folder))
    else:
        week_folder, week, year, report_date = find_week_folder()

    print(f"  Week: {week}/{year}")
    print(f"  Folder: {week_folder}")
    print()

    results    = {}   # prj_id → {found, data}
    seeds_js   = []
    missing    = []
    not_due    = []

    # Loaded up front: the main loop needs it to tell a late report from one
    # that simply is not due yet.
    history = load_history()

    # Process each project group
    for folder_name, prj_ids in FOLDER_PROJECTS.items():
        group_path = resolve_group_folder(week_folder, folder_name)
        folder_exists = group_path is not None

        for prj_id in prj_ids:
            name = PROJECT_NAMES.get(prj_id, prj_id)
            print(f"  [{prj_id}] {name}", end=' ... ')

            pdf = find_pdf(group_path, prj_id) if folder_exists else None
            if not pdf:
                reason = ('FOLDER NOT FOUND (%s)' % folder_name
                          if not folder_exists else 'PDF NOT FOUND')
                due, note = is_report_due(history, prj_id, week, year)
                suffix = f' — {note}' if note else ''
                print(f"{'' if due else 'NOT DUE — '}{reason}{suffix}")
                results[prj_id] = {'found': False,
                                   'data': {'concerns': [], 'activities': []}}
                (missing if due else not_due).append(prj_id)
                continue

            note = " (+ manual override)" if prj_id in MANUAL_OVERRIDES else ""
            print(f"OK → {os.path.basename(pdf)}{note}")
            data = extract_from_pdf(pdf, prj_id, search_dir=group_path)
            # Hand-written fields win over the extractor, but only the ones
            # actually listed — everything else stays as read from the PDF.
            if prj_id in MANUAL_OVERRIDES:
                data = {**data, **MANUAL_OVERRIDES[prj_id]}
            plan_s   = f"{data['plan']}%" if data['plan'] is not None else 'null'
            actual_s = f"{data['actual']}%" if data['actual'] is not None else 'null'
            scopes_n = len(data.get('scopes', {}))
            discs_n  = len(data.get('disciplines', {}))
            struct   = f"scopes={scopes_n}" if scopes_n else f"discs={discs_n}"
            print(f"       concerns={len(data['concerns'])}, activities={len(data['activities'])}, plan={plan_s}, actual={actual_s}, {struct}")
            results[prj_id] = {'found': True, 'data': data}
            seeds_js.append(build_seed(prj_id, week, year, data))

    print()
    print(f"  Found:   {sum(1 for r in results.values() if r['found'])} projects")
    print(f"  Missing: {len(missing)} projects: {missing}")
    if not_due:
        print(f"  Not due: {len(not_due)} projects: {not_due}")

    # Quality gate. Every project that produced a PDF used to print "OK"
    # whatever came out of it, so a pattern that stopped matching stayed
    # invisible until someone opened the dashboard - Pak Lay ran null for two
    # weeks and GMTP for five that way. Check the shape of each result against
    # what its report type always yields, and against what the project itself
    # reported last week.
    if args.dump:
        with open(args.dump, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=1, ensure_ascii=False)
        print(f"  [dump] raw results written to {args.dump}")

    report = validate_run(results, week, year, history=history)
    print()
    print(format_findings(report))

    if args.dry_run:
        print("\n  [dry-run] Skipping file updates.")
        return

    if args.strict and report['counts'][FAIL]:
        print(f"\n  [strict] {report['counts'][FAIL]} FAIL finding(s) - "
              f"stopping before any file is written.")
        sys.exit(1)

    # Update files
    print("\n  Updating dashboard...")
    update_html(week, year, seeds_js, missing, report_date=report_date)
    update_excel(week, year, results)
    if args.no_push:
        print("  [no-push] Skipping git commit/push.")
    else:
        git_push(week, year)
    # Recorded after validation so this week is never its own baseline.
    save_history(record_run(history, results, week, year))

    if not args.no_email:
        send_email(week, year, results, missing,
                   report_date=report_date)
    notify(week, year,
           sum(1 for r in results.values() if r['found']),
           len(missing))

    print("\n  Done! Dashboard URL:", DASHBOARD_URL)
    print("=" * 60)


if __name__ == '__main__':
    main()
