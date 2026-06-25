#!/usr/bin/env python3
"""
export_report.py — Gulf Energy Development Weekly Report Exporter
Reads the Gulf_Dashboard_W*.xlsx file, builds JSON, generates the Word report.

Usage:
    python export_report.py                          # auto-detect latest xlsx
    python export_report.py --week 24 --year 2026   # specific week
    python export_report.py --xlsx path/to/file.xlsx # explicit file
"""

import argparse
import glob
import json
import os
import platform
import re
import subprocess
import sys
import tempfile
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ── Excel helpers ──────────────────────────────────────────────────────────────

def find_xlsx(week: int | None, year: int | None, base_dir: str) -> str:
    """Locate the Gulf Dashboard Excel file."""
    pattern = os.path.join(base_dir, "Gulf_Dashboard_W*.xlsx")
    candidates = glob.glob(pattern)

    if not candidates:
        # Fallback: search one level up
        pattern2 = os.path.join(os.path.dirname(base_dir), "Gulf_Dashboard_W*.xlsx")
        candidates = glob.glob(pattern2)

    if not candidates:
        raise FileNotFoundError(
            f"No Gulf_Dashboard_W*.xlsx found in {base_dir}. "
            "Pass --xlsx to specify a path."
        )

    if week is not None and year is not None:
        # Filter by week/year encoded in filename
        target = f"W{year}_{week:02d}" if year else f"W{week:02d}"
        matched = [f for f in candidates if target in os.path.basename(f)]
        if matched:
            return sorted(matched)[-1]

    # Return most recently modified
    return max(candidates, key=os.path.getmtime)


def col_letter_to_index(letter: str) -> int:
    """Convert Excel column letter(s) to 0-based index."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1


def cell_value(ws, row, col):
    """Get cell value by (row, col) both 1-based."""
    return ws.cell(row=row, column=col).value


def find_header_row(ws, keywords=("project", "name", "id")):
    """Find the row index (1-based) containing column headers."""
    for row_idx in range(1, min(20, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=row_idx, column=c).value or "").lower()
                    for c in range(1, min(20, ws.max_column + 1))]
        matches = sum(1 for kw in keywords if any(kw in v for v in row_vals))
        if matches >= 2:
            return row_idx
    return 1


def map_headers(ws, header_row: int) -> dict:
    """Return {lower_header_text: col_index_1based}."""
    mapping = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            mapping[str(val).strip().lower()] = col
    return mapping


def find_col(headers: dict, *candidates) -> int | None:
    """Find first matching column from candidate header names."""
    for c in candidates:
        if c in headers:
            return headers[c]
    return None


def split_multiline(value) -> list[str]:
    """Split a cell value that may contain multiple items separated by newlines/semicolons."""
    if value is None:
        return []
    text = str(value).strip()
    if not text or text.lower() in ("-", "n/a", "none", ""):
        return []
    # Split on newlines or semicolons
    parts = re.split(r"[\n;]+", text)
    return [p.strip() for p in parts if p.strip()]


def parse_html_progress(html_path: str, week: int, year: int) -> dict:
    """Extract seeded progress data (plan/actual/scopes) from index.html for a given week/year.
    Returns {pid: {plan, actual, scopes, contract}} dict."""
    if not os.path.isfile(html_path):
        return {}

    with open(html_path, encoding="utf-8", errors="ignore") as f:
        content = f.read()

    result = {}

    # Find all seedIfEmpty('PRJ-XXX_WYY_ZZZZ', {...}) blocks
    seed_pattern = re.compile(
        r"seedIfEmpty\(\s*['\"]([^'\"]+)['\"]\s*,\s*(\{)",
        re.DOTALL,
    )

    for m in seed_pattern.finditer(content):
        key = m.group(1)  # e.g. "PRJ-002_W24_2026"

        # Parse key to extract project id, week, year
        km = re.match(r"(PRJ-\d+)_W(\d+)_(\d+)", key)
        if not km:
            continue
        pid   = km.group(1)
        k_week = int(km.group(2))
        k_year = int(km.group(3))
        if k_week != week or k_year != year:
            continue

        # Extract the JS object by counting braces
        start = m.start(2)
        depth = 0
        end = start
        for i, ch in enumerate(content[start:], start):
            if ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break

        obj_str = content[start:end]

        # Top-level plan/actual are on the first content line: "plan: 83.04, actual: 59.19,"
        top_m  = re.search(r"\bplan\s*:\s*([\d.]+)[^}]*?\bactual\s*:\s*([\d.]+)", obj_str)
        plan   = float(top_m.group(1)) if top_m else None
        actual = float(top_m.group(2)) if top_m else None

        # Extract scopes block if present
        scopes = {}
        scopes_m = re.search(r"\bscopes\s*:\s*\{", obj_str)
        if scopes_m:
            s_start = obj_str.index('{', scopes_m.start())
            depth2 = 0
            s_end = s_start
            for i, ch in enumerate(obj_str[s_start:], s_start):
                if ch == '{':
                    depth2 += 1
                elif ch == '}':
                    depth2 -= 1
                    if depth2 == 0:
                        s_end = i + 1
                        break
            scopes_str = obj_str[s_start:s_end]

            # Match only TOP-LEVEL scope entries (depth-1 inside scopes block):
            # 'Scope Name': { plan: X, actual: Y, ...}
            # Walk character by character to find entries at depth 1 only.
            pos = 1  # skip opening '{'
            scope_depth = 0
            while pos < len(scopes_str) - 1:
                ch = scopes_str[pos]
                if ch in ('"', "'") and scope_depth == 0:
                    # Find closing quote
                    q = ch
                    end_q = scopes_str.index(q, pos + 1)
                    scope_name = scopes_str[pos + 1:end_q]
                    # Find the '{' that starts this scope's value
                    brace_pos = scopes_str.index('{', end_q)
                    # Extract that inner object by brace counting
                    d = 0
                    inner_end = brace_pos
                    for j, c2 in enumerate(scopes_str[brace_pos:], brace_pos):
                        if c2 == '{': d += 1
                        elif c2 == '}':
                            d -= 1
                            if d == 0:
                                inner_end = j + 1
                                break
                    inner = scopes_str[brace_pos:inner_end]
                    # Extract plan/actual from this inner object (first occurrence)
                    pa_m = re.search(r"\bplan\s*:\s*([\d.]+)[^}]*?\bactual\s*:\s*([\d.]+)", inner)
                    if pa_m and scope_name:
                        scopes[scope_name] = {
                            "plan":   float(pa_m.group(1)),
                            "actual": float(pa_m.group(2)),
                        }
                    pos = inner_end
                    continue
                if ch == '{':
                    scope_depth += 1
                elif ch == '}':
                    scope_depth -= 1
                pos += 1

        result[pid] = {"plan": plan, "actual": actual, "scopes": scopes}

    return result


def parse_xlsx(xlsx_path: str, week: int | None, year: int | None) -> dict:
    """Parse the Excel workbook and return the JSON-compatible data dict."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Determine week/year from filename if not supplied
    basename = os.path.basename(xlsx_path)
    m = re.search(r"W(\d{4})_(\d+)", basename)
    if m:
        year  = year  or int(m.group(1))
        week  = week  or int(m.group(2))
    else:
        now = datetime.now()
        year  = year  or now.year
        week  = week  or now.isocalendar()[1]

    # Determine report date from ISO week
    report_date = datetime.fromisocalendar(year, week, 2)  # Tuesday of that week
    date_str = report_date.strftime("%Y-%m-%d")

    projects = []

    # Try each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue

        header_row = find_header_row(ws)
        headers = map_headers(ws, header_row)

        # We need at least an ID or Name column
        id_col   = find_col(headers, "id", "project id", "prj id", "project_id")
        name_col = find_col(headers, "name", "project name", "project_name", "title")
        if id_col is None and name_col is None:
            continue

        type_col      = find_col(headers, "type", "project type", "category")
        contract_col  = find_col(headers, "contract", "contract type", "contract_type")
        pdf_col       = find_col(headers, "pdf", "pdf found", "pdf_found", "pdf status")
        concerns_col  = find_col(headers, "concerns", "concern", "issues", "risks")
        activities_col = find_col(headers, "activities", "activity", "next activities",
                                  "next period", "planned activities")

        for row in range(header_row + 1, ws.max_row + 1):
            pid  = cell_value(ws, row, id_col)   if id_col   else None
            name = cell_value(ws, row, name_col) if name_col else None

            # Skip blank rows
            if pid is None and name is None:
                continue
            if str(pid or "").strip() == "" and str(name or "").strip() == "":
                continue

            pid  = str(pid or "").strip()  or f"PRJ-{row:03d}"
            name = str(name or "").strip() or "Unknown Project"

            ptype    = str(cell_value(ws, row, type_col)     or "").strip() if type_col     else ""
            contract = str(cell_value(ws, row, contract_col) or "").strip() if contract_col else ""

            # PDF found: interpret as boolean-ish
            pdf_raw = cell_value(ws, row, pdf_col) if pdf_col else None
            if pdf_raw is None:
                pdf_found = False
            elif isinstance(pdf_raw, bool):
                pdf_found = pdf_raw
            else:
                pdf_str = str(pdf_raw).strip().lower()
                pdf_found = pdf_str in ("true", "yes", "1", "found", "✓", "ok", "x")

            concerns   = split_multiline(cell_value(ws, row, concerns_col))   if concerns_col   else []
            activities = split_multiline(cell_value(ws, row, activities_col)) if activities_col else []

            projects.append({
                "id":         pid,
                "name":       name,
                "type":       ptype,
                "contract":   contract,
                "pdf_found":  pdf_found,
                "concerns":   concerns,
                "activities": activities,
                "plan":       None,
                "actual":     None,
                "scopes":     {},
            })

        if projects:
            break  # Use first sheet with data

    wb.close()

    # Merge progress data from index.html (scopes, plan/actual)
    html_path = os.path.join(os.path.dirname(os.path.dirname(xlsx_path)), "index.html")
    html_data = parse_html_progress(html_path, week, year)
    for p in projects:
        hd = html_data.get(p["id"])
        if hd:
            if p["plan"]   is None: p["plan"]   = hd["plan"]
            if p["actual"] is None: p["actual"] = hd["actual"]
            if not p["scopes"]:    p["scopes"]  = hd["scopes"]
            # Fill concerns/activities from HTML if Excel didn't have them
            if not p["concerns"]   and hd.get("concerns"):
                p["concerns"]   = hd["concerns"]
            if not p["activities"] and hd.get("activities"):
                p["activities"] = hd["activities"]

    return {
        "week":     week,
        "year":     year,
        "date":     date_str,
        "projects": projects,
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def open_file(path: str):
    """Open a file with the default application."""
    if platform.system() == "Windows":
        os.startfile(path)
    elif platform.system() == "Darwin":
        subprocess.run(["open", path])
    else:
        subprocess.run(["xdg-open", path])


def main():
    parser = argparse.ArgumentParser(
        description="Export Gulf Energy Development weekly report to Word (.docx)"
    )
    parser.add_argument("--week",  type=int, default=None, help="ISO week number")
    parser.add_argument("--year",  type=int, default=None, help="Year")
    parser.add_argument("--xlsx",  type=str, default=None, help="Path to Excel file")
    parser.add_argument("--output", type=str, default=None, help="Output .docx path")
    parser.add_argument("--no-open", action="store_true", help="Don't open the output file")
    args = parser.parse_args()

    # Resolve paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)  # one level up from automation/

    # Locate Excel
    if args.xlsx:
        xlsx_path = args.xlsx
        if not os.path.isfile(xlsx_path):
            print(f"ERROR: Excel file not found: {xlsx_path}")
            sys.exit(1)
    else:
        try:
            xlsx_path = find_xlsx(args.week, args.year, project_dir)
            print(f"Using Excel: {xlsx_path}")
        except FileNotFoundError as e:
            print(f"ERROR: {e}")
            sys.exit(1)

    # Parse Excel
    print("Parsing Excel data...")
    data = parse_xlsx(xlsx_path, args.week, args.year)
    print(f"  Week {data['week']} / {data['year']}  —  {len(data['projects'])} projects")

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        out_name = f"Gulf_Progress_Report_W{data['year']}_{data['week']:02d}.docx"
        output_path = os.path.join(project_dir, out_name)

    # Write temp JSON
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    ) as tmp:
        json.dump(data, tmp, ensure_ascii=False, indent=2)
        tmp_json = tmp.name

    try:
        js_script = os.path.join(script_dir, "generate_report.js")
        if not os.path.isfile(js_script):
            print(f"ERROR: generate_report.js not found at {js_script}")
            sys.exit(1)

        print(f"Generating Word document -> {output_path}")
        result = subprocess.run(
            ["node", js_script, tmp_json, output_path],
            capture_output=True,
            text=True,
        )

        if result.stdout:
            print(result.stdout.strip())
        if result.stderr:
            print("STDERR:", result.stderr.strip(), file=sys.stderr)

        if result.returncode != 0:
            print(f"ERROR: node exited with code {result.returncode}")
            sys.exit(result.returncode)

    finally:
        os.unlink(tmp_json)

    if not args.no_open:
        print("Opening document...")
        open_file(output_path)

    print("Done.")


if __name__ == "__main__":
    main()
